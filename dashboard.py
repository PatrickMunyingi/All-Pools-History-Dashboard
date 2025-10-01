# streamlit run app.py

# streamlit run app.py

# =========================
# Imports & Config
# =========================
import os, io, time, shutil  
import numpy as np
import pandas as pd
import altair as alt
import plotly.express as px
from datetime import datetime
import streamlit as st
import requests
from openpyxl import load_workbook




# --- Paths / sheet names ---
DATA_PATH = "all pools.xlsx"
IIS_SHEET = "IIS"

st.set_page_config(
    page_title="All Pools History Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =========================
# Global Styles & Header
# =========================
st.markdown(
    """
    <style>
        [data-testid="stHeader"] { height: 0rem; }
        [data-testid="stToolbar"] { display: none; }
        @keyframes fadeInBounce {
            0% {opacity: 0; transform: translateY(-20px);}
            50% {opacity: 0.5; transform: translateY(5px);}
            100% {opacity: 1; transform: translateY(0);}
        }
        .animated-title {
            text-align: center;
            color: #1E90FF;
            font-size: 40px;
            font-weight: bold;
            animation: fadeInBounce 1.5s ease-out;
        }
    </style>
    """,
    unsafe_allow_html=True
)
st.markdown("<h1 class='animated-title'>ALL POOLS HISTORY DASHBOARD</h1>", unsafe_allow_html=True)
first_of_month = datetime.today().replace(day=1).strftime("%B %d, %Y")
st.markdown(f"** Data as of {first_of_month}**")




# =========================
# Helpers
# =========================
def sort_pools(pool_list):
    """Natural sort like 1, 2, 3, 10, 10A, 10B."""
    return sorted(pool_list, key=lambda x: (int(''.join(filter(str.isdigit, str(x))) or 0), str(x)))

@st.cache_data
def load_data_sov():
    df = pd.read_excel(DATA_PATH, sheet_name="SOV&REPLICA")
    if "Policy ID" in df.columns:
        df.set_index("Policy ID", inplace=True)
    numeric_cols = ['Premium', 'Attachment', 'Exhaustion', 'Coverage', 'Claims']
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df['Pool'] = df['Pool'].astype(str)
    df['Master Pool'] = df['Pool'].str.extract(r'(\d+)')
    df['Master Pool'] = df['Master Pool'].fillna(df['Pool'])
    return df

@st.cache_data
def load_data_iis():
    return pd.read_excel(DATA_PATH, sheet_name=IIS_SHEET)


def backup_then_replace_iis_sheet(df: pd.DataFrame, xlsx_path: str, sheet_name: str = IIS_SHEET):
    """Backs up workbook, then replaces only the IIS sheet with df."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")
    # backup
    ts = time.strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.splitext(xlsx_path)[0] + f"_BACKUP_{ts}.xlsx"
    shutil.copy2(xlsx_path, backup_path)
    # remove / create target sheet
    wb = load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    wb.create_sheet(sheet_name)
    wb.save(xlsx_path)
    # write replacement IIS
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return backup_path

def pick(df, *cands):
    """Return the first existing column matching any candidate (case/space-insensitive)."""
    norm = {str(c).strip().lower().replace(" ", ""): c for c in df.columns}
    for cand in cands:
        key = str(cand).strip().lower().replace(" ", "")
        if key in norm: return norm[key]
    return None

# =========================
# App Selector
# =========================
Business_Types = st.selectbox("Choose Business Type", ("", "SOVEREIGN BUSINESS", "IIS"))

# =============================================================================
# SOVEREIGN BUSINESS
# =============================================================================
if Business_Types == "SOVEREIGN BUSINESS":
    df = load_data_sov()
    def normalize_columns(df):
        df = df.copy()
        df.columns = (
            df.columns.astype(str)
            .str.replace('\u00A0', ' ', regex=False)   # NBSP -> space
            .str.replace(r'\s+', ' ', regex=True)      # collapse spaces
            .str.strip()
        )
        # Alias common variants to the canonical name
        aliases = {
            "policyholder": "Policy Holder",
            "policy_holder": "Policy Holder",
            "policy holder": "Policy Holder",
        }
        lower_map = {c.lower(): c for c in df.columns}
        for k, target in aliases.items():
            if k in lower_map and target not in df.columns:
                df.rename(columns={lower_map[k]: target}, inplace=True)
        return df

    df = normalize_columns(df)
    premium_payers = [c for c in df.columns if str(c).startswith("Premium Financed by")]

    # Resolve column names robustly
    COL_POOL        = 'Pool'
    COL_MASTERPOOL  = 'Master Pool'
    COL_POLICYTYPE  = pick(df, 'Policy Type', 'PolicyType')
    COL_COUNTRY     = pick(df, 'Country')
    COL_REGION      = pick(df, 'Region')
    COL_PERIL       = pick(df, 'Peril')
    COL_CROPTYPE    = pick(df, 'Crop Type', 'CropType')
    COL_POLICYHOLDER= pick(df, 'Policy Holder', 'Policy_Holder', 'PolicyHolder', 'Policy holder')

    # ---- Sidebar Filters ----
    with st.sidebar.expander("Filters", expanded=True):
        show_sub_pools   = st.checkbox("Show Sub-Pools (like 10A, 10B)", value=False)
        pool_column      = COL_POOL if show_sub_pools else COL_MASTERPOOL

        sorted_pool_options = sort_pools(df[pool_column].astype(str).unique())
        select_all_pools  = st.checkbox("Select All Pools", value=True)
        pool = st.multiselect("Select Pool:", options=sorted_pool_options,
                              default=sorted_pool_options if select_all_pools else [])

        select_all_policy_types = st.checkbox("Select All Policy Types", value=True)
        policy_type = st.multiselect("Policy Type:",
                                     options=df[COL_POLICYTYPE].dropna().unique().tolist() if COL_POLICYTYPE else [],
                                     default=(df[COL_POLICYTYPE].dropna().unique().tolist()
                                              if select_all_policy_types and COL_POLICYTYPE else []))

        select_all_countries = st.checkbox("Select All Countries", value=True)
        country = st.multiselect("Country:",
                                 options=df[COL_COUNTRY].dropna().unique().tolist() if COL_COUNTRY else [],
                                 default=(df[COL_COUNTRY].dropna().unique().tolist()
                                          if select_all_countries and COL_COUNTRY else []))

        select_all_regions = st.checkbox("Select All Regions", value=True)
        region = st.multiselect("Region:",
                                options=df[COL_REGION].dropna().unique().tolist() if COL_REGION else [],
                                default=(df[COL_REGION].dropna().unique().tolist()
                                         if select_all_regions and COL_REGION else []))

        select_all_peril = st.checkbox("Select All Perils", value=True)
        peril = st.multiselect("Peril:",
                               options=df[COL_PERIL].dropna().unique().tolist() if COL_PERIL else [],
                               default=(df[COL_PERIL].dropna().unique().tolist()
                                        if select_all_peril and COL_PERIL else []))

        #  fix: use select_all_crop_types here (was select_all_peril)
        select_all_crop_types = st.checkbox("Select All Crop Types", value=True)
        crop_type = st.multiselect("Crop Type:",
                                   options=df[COL_CROPTYPE].dropna().unique().tolist() if COL_CROPTYPE else [],
                                   default=(df[COL_CROPTYPE].dropna().unique().tolist()
                                            if select_all_crop_types and COL_CROPTYPE else []))

        select_all_policy_holders = st.checkbox("Select All Policy Holders", value=True)
        policy_options = df[COL_POLICYHOLDER].dropna().astype(str).unique().tolist() if COL_POLICYHOLDER else []
        policy_holders = st.multiselect("Policy Holder:", options=policy_options,
                                        default=policy_options if select_all_policy_holders else [])

    # ---- Safe masks (no KeyErrors even if a column is missing) ----
    mask = (
        df[pool_column].isin(pool)
        & (df[COL_POLICYTYPE].isin(policy_type) if COL_POLICYTYPE else True)
        & (df[COL_COUNTRY].isin(country) if COL_COUNTRY else True)
        & (df[COL_PERIL].isin(peril) if COL_PERIL else True)
        & (df[COL_REGION].isin(region) if COL_REGION else True)
        & (df[COL_CROPTYPE].isin(crop_type) if COL_CROPTYPE else True)
        & (df[COL_POLICYHOLDER].astype(str).isin(policy_holders) if COL_POLICYHOLDER else True)
    )

    df_selection = df[mask]
    num_policies = len(df_selection)

    # Optional: let the user know if a key column is missing for this Business Type
    missing = [name for name, col in {
        "Policy Type": COL_POLICYTYPE,
        "Country": COL_COUNTRY,
        "Region": COL_REGION,
        "Peril": COL_PERIL,
        "Crop Type": COL_CROPTYPE,
        "Policy Holder": COL_POLICYHOLDER,
    }.items() if not col]
    if missing:
        st.info(f"Note: The dataset lacks these fields: {', '.join(missing)}.")

    num_policies = len(df_selection)
    
    # ---- View Selection ----
    option = st.selectbox(
        "What would you like to view?",
        ("","Insurance Footprint Map", "Premium and Country Basic Information", "Premium Financing and Tracker", "Claim Settlement History")
    )
    #----------------------------
    #Section 0: Our FootPrint
    #----------------------------
    #Create Chrolopleth maps
    if option == "Insurance Footprint Map":
            st.markdown(
                "<span style='font-weight:bold; font-size:18px;'>🌎Visualizing regional activity to track footprint and insurance presence across Africa</span>",
                unsafe_allow_html=True
            )
        
            map_metric = st.radio(  # <-- Line 247
                "Select metric:",
                ["Claims", "Premium", "Loss Ratio", "Coverage", "Number of Policies"],
                horizontal=True
            )


             # --- Base sums per country
            country_stats = (
                 df_selection.groupby("Country", as_index=False)[["Claims", "Premium", "Coverage"]]
                 .sum()
             )

             # --- Policy counts per country (row count). If you have a unique policy-id column, use nunique on it.
            # counts = df_selection.groupby("Country")["Policy Ref"].nunique().reset_index(name="Number of Policies")
            counts = df_selection.groupby("Country").size().reset_index(name="Number of Policies")
            country_stats = country_stats.merge(counts, on="Country", how="left").fillna({"Number of Policies": 0})
            # --- Loss Ratio (avoid divide-by-zero)
            country_stats["Loss Ratio"] = np.where(
                country_stats["Premium"] > 0,
                (country_stats["Claims"] / country_stats["Premium"]) * 100,
                np.nan
            )
            color_scale = {
                "Claims": "Reds",
                "Premium": "Blues",
                "Loss Ratio": "Oranges",
                "Coverage": "Greens",
                "Number of Policies": "Purples",
            }
            title_map = {
                "Claims": "Total Claims by Country",
                "Premium": "Total Premium by Country",
                "Loss Ratio": "Loss Ratio (%) by Country",
                "Coverage": "Total Coverage by Country",
                "Number of Policies": "Number of Policies by Country",
            }
            if not country_stats.empty:
                fig_map = px.choropleth(
                    country_stats,
                    locations="Country",
                    locationmode="country names",
                    color=map_metric,
                    hover_name="Country",
                    color_continuous_scale=color_scale[map_metric],
                    title=f"🌍 {title_map[map_metric]}",
                    template="plotly_white",
                    scope="africa",
                )
                # Bigger, cleaner map
                fig_map.update_geos(
                    showcountries=True, countrycolor="#1f1f1f",
                    showcoastlines=False, showland=True, landcolor="rgba(240,240,240,0.6)",
                    fitbounds="locations"
                )
                fig_map.update_layout(
                    height=1000,width=1000,
                    margin=dict(l=0, r=0, t=56, b=0),
                    coloraxis_colorbar=dict(len=0.9, thickness=14)
                )
                st.plotly_chart(fig_map, use_container_width=True, config={"displayModeBar": False})
            else:
                st.info("No country-level data available for the selected metric.")
            with st.expander("View Country-Level Table"):
                 st.dataframe(
                     country_stats.sort_values(map_metric, ascending=False, na_position="last"),
                     use_container_width=True
                 )



    # ---------------------------
    # SECTION 1: Premium & Country
    # ---------------------------
    if option == "Premium and Country Basic Information":
        total_premium = df_selection['Premium'].sum()
        total_claims = df_selection['Claims'].sum()
        total_coverage = df_selection['Coverage'].sum()
        loss_ratio = (total_claims / total_premium) * 100 if total_premium > 0 else 0

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total Premium", f"US ${total_premium:,.0f}")
        c2.metric("Loss Ratio", f"{loss_ratio:.2f}%")
        c3.metric("Coverage", f"US ${total_coverage:,.0f}")
        c4.metric("Claims", f"US ${total_claims:,.0f}")
        c5.metric("Number of Policies", f"{num_policies}")

        col1, col2, col3 = st.columns(3)

        # Trend by pool
        with col1:
            if not df_selection.empty:
                trend_metric = st.radio("Select Metric", ["Premium", "Coverage"], horizontal=True)
                pool_trend = df_selection.groupby(pool_column)[trend_metric].sum().reset_index()
                pool_trend[pool_column] = pool_trend[pool_column].astype(str)
                pool_trend["__num"] = pool_trend[pool_column].str.extract(r"(\d+)").astype(int)
                pool_trend["__has_suffix"] = pool_trend[pool_column].str.contains(r"[A-Za-z]")
                ordered_labels = pool_trend.sort_values(["__has_suffix", "__num"])[pool_column].tolist()
                pool_trend[pool_column] = pd.Categorical(pool_trend[pool_column], categories=ordered_labels, ordered=True)
                fig1 = px.line(
                    pool_trend.sort_values([pool_column]), x=pool_column, y=trend_metric, markers=True,
                    title=f'Yearly {trend_metric}s Over Time', template='plotly_white',
                    category_orders={pool_column: ordered_labels},labels={pool_column:"Pool", trend_metric: trend_metric}
                )
                st.plotly_chart(fig1, use_container_width=True)

        # Country count
        with col2:
            country_count = df_selection['Country'].value_counts().reset_index()
            country_count.columns = ['Country', 'Number of Policies']
            fig2 = px.bar(country_count, x='Number of Policies', y='Country', orientation='h', title="Policy Count by Country",
                          template='plotly_white')
            fig2.update_traces(texttemplate='%{x:,.0f}', textposition='outside')
            st.plotly_chart(fig2, use_container_width=True)

        # Policy type distribution
        with col3:
            policy_type_counts = df_selection['Policy Type'].value_counts().reset_index()
            policy_type_counts.columns = ['Policy Type', 'Count']
            fig3 = px.pie(policy_type_counts, names='Policy Type', values='Count', hole=0.6, title="Policy Type Distribution")
            st.plotly_chart(fig3, use_container_width=True)

        # Table (pretty formatting, guarding missing cols)
        st.markdown("### Filtered Data")
        export_df = df_selection.copy()
        if 'Rate-On-Line' in export_df.columns:
            export_df['Rate-On-Line'] = pd.to_numeric(export_df['Rate-On-Line'], errors='coerce') \
                .apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
        if 'Ceding %' in export_df.columns:
            export_df['Ceding %'] = pd.to_numeric(export_df['Ceding %'], errors='coerce') \
                .apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
        for col in export_df.columns:
            if col not in ['Rate-On-Line', 'Ceding %', 'Premium Loading'] and pd.api.types.is_numeric_dtype(export_df[col]):
                export_df[col] = export_df[col].apply(lambda x: f"{x:,.0f}")
        st.dataframe(export_df, use_container_width=True)

    # ---------------------------
    # SECTION 2: Premium Financing
    # ---------------------------
    elif option == "Premium Financing and Tracker":
        mapping = {col: col.replace("Premium Financed by ", "") for col in premium_payers}
        st.markdown(
            "### Select Premium Payers",
            help="Note: Pools 1–5 had no premium financing; it began at Pool 6 (2019/2020)."
        )
        select_all = st.checkbox("Select All Premium Payers", value=True)
        picked_display = st.multiselect("Premium Payers", mapping.values(), default=mapping.values() if select_all else [])
        picked_cols = [k for k, v in mapping.items() if v in picked_display]

        if not picked_cols:
            df_pf = df_selection
            total_prem = df_pf.get("Premium", pd.Series(dtype=float)).sum()
        else:
            mask = df_selection[picked_cols].fillna(0).sum(axis=1) > 0
            df_pf = df_selection[mask]
            total_prem = df_pf[picked_cols].sum().sum()

        total_claims = df_selection.get("Claims", pd.Series(dtype=float)).sum()
        total_cov = df_selection.get("Coverage", pd.Series(dtype=float)).sum()
        loss_ratio = (total_claims / total_prem) * 100 if total_prem > 0 else 0

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total Premium (from Payers)", f"US ${total_prem:,.0f}")
        c2.metric("Loss Ratio", f"{loss_ratio:.2f}%")
        c3.metric("Coverage", f"US ${total_cov:,.0f}")
        c4.metric("Claims", f"US ${total_claims:,.0f}")
        c5.metric("Number of Policies", f"{num_policies}")

        chart_view = st.radio(
                                        "Chart Type",
                                        [
                                            "Donor-Style Summary",            # existing
                                            "Stacked by Pool",                # existing
                                              
                                        ],
                                        horizontal=True
                                    )

        distinct_colors = [
            "#e6194B","#3cb44b","#ffe119","#4363d8","#f58231","#911eb4","#46f0f0",
            "#f032e6","#bcf60c","#fabebe","#008080","#e6beff","#9a6324","#fffac8",
            "#800000","#aaffc3","#808000","#ffd8b1","#000075","#808080"
        ]

        if picked_cols:
            if chart_view == "Donor-Style Summary":
                s = df_pf[picked_cols].sum().reset_index()
                s.columns = ["Payer", "Amount"]
                s["Payer"] = s["Payer"].map(mapping)
                total = s["Amount"].sum() or 1
                s["%"] = s["Amount"] / total * 100
                s["Label"] = s["%"].apply(lambda x: f"{x:.2f}%") + "<br>" + s["Amount"].apply(lambda x: f"${x/1e6:.2f}m")
                fig = px.bar(
                    s, x="Payer", y="Amount", text="Label", color="Payer",
                    title="Premium Contribution by Financiers", template="plotly_white",
                    color_discrete_sequence=distinct_colors
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                melted = df_pf[[pool_column] + picked_cols].melt(id_vars=pool_column, var_name="Payer", value_name="Amount")
                melted["Payer"] = melted["Payer"].map(mapping)
                all_pools = sort_pools(df[pool_column].dropna().astype(str).unique().tolist()) if pool_column in df.columns else []
                all_payers = melted["Payer"].unique().tolist()
                full = pd.MultiIndex.from_product([all_pools, all_payers], names=[pool_column, "Payer"]).to_frame(index=False)
                grouped_actual = melted.groupby([pool_column, "Payer"], as_index=False)["Amount"].sum()
                grouped = full.merge(grouped_actual, on=[pool_column, "Payer"], how="left").fillna(0)
                fig = px.bar(
                    grouped, x=pool_column, y="Amount", color="Payer",
                    title="Premium Payers per Pool (Stacked)", barmode="stack",
                    text_auto=".2s", template="plotly_white", color_discrete_sequence=distinct_colors
                )
                fig.update_layout(xaxis={"categoryorder": "array", "categoryarray": all_pools})
                st.plotly_chart(fig, use_container_width=True)

                

            st.markdown("#### Filtered Financing Data")
            export_df = df_pf.copy()
            if 'Rate-On-Line' in export_df.columns:
                export_df['Rate-On-Line'] = pd.to_numeric(export_df['Rate-On-Line'], errors='coerce') \
                    .apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
            if 'Ceding %' in export_df.columns:
                export_df['Ceding %'] = pd.to_numeric(export_df['Ceding %'], errors='coerce') \
                    .apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
            for col in export_df.columns:
                if col not in ['Rate-On-Line', 'Ceding %', 'Premium Loading'] and pd.api.types.is_numeric_dtype(export_df[col]):
                    export_df[col] = export_df[col].apply(lambda x: f"{x:,.0f}")
            st.dataframe(export_df, use_container_width=True)

    # ---------------------------
    # SECTION 3: Claims
    # ---------------------------
    elif option == "Claim Settlement History":
        st.subheader("Claim Settlement Overview")
        total_claims = df_selection['Claims'].sum()
        num_claims = df_selection[df_selection["Claims"] > 0].shape[0]
        avg_claim = total_claims / num_claims if num_claims > 0 else 0
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Claims", f"US ${total_claims:,.0f}")
        c2.metric("Number of Policies", f"{num_policies}")
        c3.metric("Number of Claims", f"{num_claims}")
        c4.metric("Avg Claim (per Claim)", f"US ${avg_claim:,.0f}")

        sorted_all_pools = sort_pools(df[pool_column].unique())
        claims_by_pool = df_selection.groupby(pool_column, as_index=False)["Claims"].sum()
        claims_by_pool = pd.DataFrame({pool_column: sorted_all_pools}).merge(claims_by_pool, on=pool_column, how="left").fillna(0)

        col1, col2, col3 = st.columns(3)
        with col1:
            top_pools = df_selection.groupby(pool_column)["Claims"].sum().sort_values(ascending=False).head(10).reset_index()
            fig1 = px.bar(
                top_pools, x="Claims", y=pool_column, orientation="h",
                title="💰 Top 10 Pools by Claims Paid", text="Claims",
                template="plotly_white", color="Claims",labels={pool_column:"Pool","Claims": "Claims(USD)"}
            )
            fig1.update_traces(texttemplate='$%{x:,.0f}', textposition='outside')
            st.plotly_chart(fig1)
        with col2:
            if not df_selection.empty:
                import re
                import numpy as np
                import plotly.graph_objects as go

                # --- Clean numerics (handles "47,000,000", "$60,123,456", etc.)
                for col in ["Premium", "Claims"]:
                    if df_selection[col].dtype == "O":
                        df_selection[col] = (
                            df_selection[col]
                            .astype(str)
                            .str.replace(r"[^\d\.\-]", "", regex=True)  # remove $, commas, spaces
                        )
                    df_selection[col] = pd.to_numeric(df_selection[col], errors="coerce").fillna(0)

                # --- Aggregate per pool
                pool_trend = (
                    df_selection.groupby(pool_column)[["Premium", "Claims"]]
                    .sum()
                    .reset_index()
                )
                pool_trend[pool_column] = pool_trend[pool_column].astype(str)

                # --- Order: 10, 10A, 10B…
                pool_trend["__num"] = pool_trend[pool_column].str.extract(r"(\d+)").astype(float)
                pool_trend["__has_suffix"] = pool_trend[pool_column].str.contains(r"[A-Za-z]")
                ordered_labels = (
                    pool_trend.sort_values(["__num", "__has_suffix"])[pool_column].tolist()
                )
                pool_trend[pool_column] = pd.Categorical(
                    pool_trend[pool_column], categories=ordered_labels, ordered=True
                )
                pool_trend = pool_trend.sort_values(pool_column).copy()

                # --- Pretty labels
                def short(x):
                    if x is None or np.isnan(x): return "—"
                    ax = abs(x)
                    if ax >= 1e9:  return f"{x/1e9:.1f}B"
                    if ax >= 1e6:  return f"{x/1e6:.0f}M"
                    if ax >= 1e3:  return f"{x/1e3:.0f}k"
                    return f"{x:,.0f}"

                # --- Single-axis combo: bars = Premium, line = Claims
                fig2 = go.Figure()

                fig2.add_bar(
                    x=pool_trend[pool_column],
                    y=pool_trend["Premium"],
                    name="Premium",
                    marker_color="#1f77b4",
                    text=[short(v) for v in pool_trend["Premium"]],
                    textposition="outside",
                    hovertemplate="<b>%{x}</b><br>Premium: %{y:,.0f}<extra></extra>",
                )

                fig2.add_trace(
                    go.Scatter(
                        x=pool_trend[pool_column],
                        y=pool_trend["Claims"],
                        name="Claims",
                        mode="lines+markers",
                        marker=dict(size=8),
                        line=dict(width=3, color="firebrick"),
                        hovertemplate="<b>%{x}</b><br>Claims: %{y:,.0f}<extra></extra>",
                    )
                )

                # --- Layout (shared y-axis for truthful comparison)
                ymax = float(max(pool_trend["Premium"].max(), pool_trend["Claims"].max()))
                fig2.update_layout(
                    template="plotly_white",
                    title="Premium & Claims by Pool",
                    xaxis_title="Pool",
                    yaxis_title="USD",
                    yaxis=dict(range=[0, ymax * 1.15]),
                    legend=dict(orientation="h", y=1.1, x=0),
                    margin=dict(t=60, b=40),
                )

                st.plotly_chart(fig2, use_container_width=True)

               

        with col3:
            pool_summary = df_selection.groupby(pool_column).agg({'Claims':'sum','Premium':'sum'}).reset_index()
            pool_summary["Loss Ratio"] = (pool_summary["Claims"]/pool_summary["Premium"]) * 100
            top_loss = pool_summary[pool_summary["Premium"]>0].sort_values("Loss Ratio", ascending=False).head(10)
            fig3 = px.bar(
                top_loss, x=pool_column, y="Loss Ratio",
                title=" Pools with Highest Loss Ratios", text="Loss Ratio",
                template="plotly_white", color='Loss Ratio',labels={pool_column:"Pool","Loss Ratio": "Loss Ratio (%)"}
            )
            fig3.update_traces(texttemplate='%{y:.1f}%', textposition='outside')
            fig3.update_layout(yaxis_title="Loss Ratio (%)")
            st.plotly_chart(fig3, use_container_width=True)

        

        # Table (pretty)
        st.markdown("#### Filtered Claim Data")
        export_df = df_selection.copy()
        if 'Rate-On-Line' in export_df.columns:
            export_df['Rate-On-Line'] = pd.to_numeric(export_df['Rate-On-Line'], errors='coerce') \
                .apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
        if 'Ceding %' in export_df.columns:
            export_df['Ceding %'] = pd.to_numeric(export_df['Ceding %'], errors='coerce') \
                .apply(lambda x: f"{x:.2%}" if pd.notna(x) else "")
        for col in export_df.columns:
            if col not in ['Rate-On-Line', 'Ceding %', 'Premium Loading'] and pd.api.types.is_numeric_dtype(export_df[col]):
                export_df[col] = export_df[col].apply(lambda x: f"{x:,.0f}")
        st.dataframe(export_df, use_container_width=True)

# =============================================================================
# IIS
# =============================================================================
# -----------------------------------------
# IIS — aligned with Sovereign layout
# -----------------------------------------
if Business_Types == "IIS":

    # ---------- helpers ----------
    @st.cache_data
    def load_data_iis():
        return pd.read_excel(DATA_PATH, sheet_name="IIS")

    def pick(df, *cands):
        """Return the first existing column matching any candidate (case/space-insensitive)."""
        norm = {str(c).strip().lower().replace(" ", ""): c for c in df.columns}
        for cand in cands:
            key = str(cand).strip().lower().replace(" ", "")
            if key in norm:
                return norm[key]
        return None

    # In-memory editable copy (used by Edit IIS data)
    if "iis_df" not in st.session_state:
        st.session_state.iis_df = load_data_iis().copy()

    # Sub-section selector (same placement/pattern as Sovereign)
    iis_option = st.selectbox(
        "What would you like to view?",
        ("", "Summary", "Disaster Finder", "Auto-Analysis", "Edit IIS data")
    )

    # ========= SUMMARY (same feel as Sovereign) =========
    if iis_option == "Summary":
        df = st.session_state.iis_df.copy()

        # Robust column detection
        col_country   = pick(df, "Country")
        col_startdate = pick(df, "Start Date", "StartDate")
        col_arc       = pick(df, "ARC Net Premium", "ARCNetPremium")
        col_fac       = pick(df, "Facultative Reinsurance Premium", "FacRePremium")
        col_payout    = pick(df, "Total Payout ($)", "TotalPayout")
        col_partner   = pick(df, "Other Key Partners", "Partner")
        col_programme = pick(df, "Programme Name")

        # Coerce types
        for c in [col_arc, col_fac, col_payout]:
            if c and c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")
        if col_startdate:
            df[col_startdate] = pd.to_datetime(df[col_startdate], errors="coerce")

        # Sidebar (only for this sub-section)
        with st.sidebar.expander("Filters", expanded=True):
            if col_startdate:
                year_list = sorted(df[col_startdate].dt.year.dropna().unique())
                sel_all_years = st.checkbox("Select All Years", value=True)
                sel_years = st.multiselect("Select Year", options=year_list,
                                           default=year_list if sel_all_years else [])
            else:
                sel_years = []
                st.caption("No Start Date column found; year filter disabled.")

            if col_country:
                countries = df[col_country].dropna().unique()
                sel_all_countries = st.checkbox("Select All Countries", value=True)
                sel_country = st.multiselect("Select Country", options=countries,
                                             default=countries if sel_all_countries else [])
            else:
                sel_country = []
                st.caption("No Country column found; country filter disabled.")

            if col_partner:
                partners = df[col_partner].dropna().unique()
                sel_all_partners = st.checkbox("Select All Partners", value=True)
                sel_partner = st.multiselect("Select Partner", options=partners,
                                             default=partners if sel_all_partners else [])
            else:
                sel_partner = []
                st.caption("No Partner column found; partner filter disabled.")

        # Apply filters
        mask = pd.Series(True, index=df.index)
        if col_startdate and sel_years:
            mask &= df[col_startdate].dt.year.isin(sel_years)
        if col_country and sel_country:
            mask &= df[col_country].isin(sel_country)
        if col_partner and sel_partner:
            mask &= df[col_partner].isin(sel_partner)
        filtered_df = df[mask].copy()

        # KPIs
        total_arc = filtered_df[col_arc].sum() if col_arc else 0.0
        total_fac = filtered_df[col_fac].sum() if col_fac else 0.0
        total_payout = filtered_df[col_payout].sum() if col_payout else 0.0
        denom = (total_arc + total_fac) if (total_arc + total_fac) > 0 else 1
        claims_ratio = total_payout / denom
        n_programmes = filtered_df[col_programme].nunique() if col_programme else 0

        st.markdown("## Inclusive Insurance Business (IIS) Dashboard")
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("ARC Premium", f"${total_arc:,.0f}")
        k2.metric("Facultative Premium", f"${total_fac:,.0f}")
        k3.metric("Total Payout", f"${total_payout:,.0f}")
        k4.metric("Claims Ratio", f"{claims_ratio:.2%}")
        k5.metric("Programmes", n_programmes)

        # Charts + table (mirrors Sovereign’s pattern)
        if all([col_country, col_arc, col_fac, col_payout]):
            country_agg = (
                filtered_df.groupby(col_country)[[col_arc, col_fac, col_payout]]
                .sum().reset_index().rename(columns={
                    col_country: "Country",
                    col_arc: "ARCNetPremium",
                    col_fac: "FacRePremium",
                    col_payout: "TotalPayout"
                })
            )

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("### Premiums vs Payouts by Country")
                fig1 = px.bar(
                    country_agg, x="Country",
                    y=["ARCNetPremium", "FacRePremium", "TotalPayout"],
                    barmode="group", template="plotly_white"
                )
                st.plotly_chart(fig1, use_container_width=True)

            #with c2:
                # Simple map (optional but aligned with your Sovereign choropleth style)
                st.markdown("### Country Map")
                denom2 = (country_agg["ARCNetPremium"] + country_agg["FacRePremium"]).replace(0, pd.NA)
                country_agg["Claims Ratio (%)"] = (country_agg["TotalPayout"] / denom2) * 100
                map_metric = st.radio(
                    "Metric", ["TotalPayout", "ARCNetPremium", "FacRePremium", "Claims Ratio (%)"],
                    horizontal=True
                )
                fig_map = px.choropleth(
                    country_agg, locations="Country", locationmode="country names",
                    color=map_metric,
                    color_continuous_scale="Blues" if map_metric != "Claims Ratio (%)" else "Oranges",
                    template="plotly_white"
                )
                fig_map.update_geos(showcountries=True, showcoastlines=True, fitbounds="locations")
                fig_map.update_layout(margin=dict(l=0, r=0, t=40, b=0))
                st.plotly_chart(fig_map, use_container_width=True)

            st.markdown("### Country Summary Table")
            st.dataframe(country_agg)
            st.download_button(
                "Download Summary CSV",
                data=country_agg.to_csv(index=False).encode("utf-8"),
                file_name="iis_country_summary.csv",
                mime="text/csv"
            )
        else:
            st.info("Country-level fields not found to build the summary table.")

    # ========= DISASTER FINDER (sidebar only when selected) =========
    elif iis_option == "Disaster Finder":
        st.title("ReliefWeb Explorer 🌍")  # keep just one emoji here

        with st.sidebar.expander("Filters", expanded=True):
            country = st.text_input("Country (leave blank for all)", "")
            disaster_type = st.text_input("Disaster Type (e.g., flood, drought)", "")
            start_date = st.date_input("Start Date", datetime(1990, 1, 1))
            end_date = st.date_input("End Date", datetime.today())
            limit = st.slider("Number of results", 10, 100, 50)

        tab1, tab2 = st.tabs(["Disasters", "Reports"])

        with tab1:
            try:
                params = {"appname":"reliefweb-explorer","limit":limit,"profile":"list","sort[]":"date.created:desc"}
                if country:
                    params["filter[field]"] = "country"
                    params["filter[value]"] = country.lower().strip()
                resp = requests.get("https://api.reliefweb.int/v1/disasters", params=params, timeout=20)
                resp.raise_for_status()
                data = resp.json().get("data", [])
                results = []
                for d in data:
                    f = d["fields"]
                    date_str = f["date"]["created"][:10]
                    dt = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if disaster_type:
                        types = [t["name"].lower() for t in f.get("type", [])]
                        if disaster_type.lower() not in types:
                            continue
                    if not (start_date <= dt <= end_date):
                        continue
                    results.append({
                        "Name": f["name"],
                        "Type": ", ".join(t["name"] for t in f.get("type", [])),
                        "Country": ", ".join(c["name"] for c in f.get("country", [])),
                        "Date": date_str,
                        "URL": f["url"]
                    })
                if not results:
                    st.info("No disasters match the filters.")
                else:
                    ddf = pd.DataFrame(results)
                    st.dataframe(ddf, use_container_width=True)
                    st.download_button(
                        "Download Disasters CSV",
                        data=ddf.to_csv(index=False).encode("utf-8"),
                        file_name="reliefweb_disasters.csv",
                        mime="text/csv"
                    )
            except Exception as e:
                st.error(f"Failed to fetch disasters: {e}")

        with tab2:
            filters = []
            if country:
                filters.append({"field":"country","value":country.lower().strip()})
            filters.append({"field":"date.created","range":
                            {"from":start_date.strftime("%Y-%m-%d"),
                             "to":end_date.strftime("%Y-%m-%d")}})
            payload = {"limit":limit,"profile":"lite","filter":{"conditions":filters},
                       "sort":[{"field":"date.created","direction":"desc"}]}
            try:
                r = requests.post("https://api.reliefweb.int/v1/reports", json=payload,
                                  params={"appname":"reliefweb-explorer"}, timeout=20)
                r.raise_for_status()
                reports = r.json().get("data", [])
                if not reports:
                    st.info("No reports found.")
                else:
                    rows = [{"Title":x["fields"]["title"],
                             "Date":x["fields"]["date"]["created"][:10],
                             "Source":", ".join([s["name"] for s in x["fields"].get("source", [])]),
                             "URL":x["fields"]["url"]} for x in reports]
                    ddf = pd.DataFrame(rows)
                    st.dataframe(ddf, use_container_width=True)
                    st.download_button(
                        "Download Reports CSV",
                        data=ddf.to_csv(index=False).encode("utf-8"),
                        file_name="reliefweb_reports.csv",
                        mime="text/csv"
                    )
            except Exception as e:
                st.error(f"Failed to fetch reports: {e}")

    # ========= AUTO-ANALYSIS (sidebar only when selected) =========
    elif iis_option == "Auto-Analysis":

        # Your existing Seasonal Data Explorer block can be kept as-is.
        # Just ensure its controls live under st.sidebar so they only appear here.
        # (No extra emojis needed.)
        # ---- place your Auto-Analysis code here ----
    
        import pandas as pd
        import numpy as np
        from datetime import datetime
        import altair as alt

       
        st.title("Seasonal Data Explorer")

        st.markdown(
            """
            **Views**: Global • By Season • Summary  
            **Aggregation**: Sum • Mean • Raw (daily)  
            - Optional: Interpolate to **daily inside season months** (breaks off-season gaps)  
            - Optional: **Rolling average** overlay (days or seasons)
            """
        )

        # ---------------------------
        # Sidebar Controls
        # ---------------------------
        with st.sidebar:
            st.header("Season window")
            start_month = st.selectbox(
                "Start month",
                options=list(range(1, 13)),
                format_func=lambda m: datetime(2000, m, 1).strftime("%b"),
                index=9  # Oct
            )
            end_month = st.selectbox(
                "End month",
                options=list(range(1, 13)),
                format_func=lambda m: datetime(2000, m, 1).strftime("%b"),
                index=3  # Apr
            )
            st.caption("If start > end, the season crosses calendar years (e.g., Oct→Apr).")

            st.header("Upload data")
            file = st.file_uploader("CSV or Excel", type=["csv", "xlsx", "xls"])

        # ---------------------------
        # Helpers
        # ---------------------------
        def load_table(file):
            if file is None: return None
            name = file.name.lower()
            return pd.read_csv(file) if name.endswith(".csv") else pd.read_excel(file)

        def coerce_datetime(series: pd.Series):
            try:
                return pd.to_datetime(series, errors="coerce", dayfirst=False)
            except Exception:
                return pd.to_datetime(series, errors="coerce")

        def assign_season_year(d: pd.Timestamp, start_m: int, end_m: int):
            if pd.isna(d): return np.nan
            m, y = d.month, d.year
            if start_m > end_m:  # cross-year (e.g., Oct–Apr)
                if m >= start_m:      return y
                elif m <= end_m:      return y - 1
                else:                 return np.nan
            else:
                return y if start_m <= m <= end_m else np.nan

        def season_label_from_year(y: int, start_m: int, end_m: int):
            mon = lambda m: datetime(2000, m, 1).strftime("%b")
            return f"{mon(start_m)}-{mon(end_m)} {y}/{y+1}" if start_m > end_m else f"{mon(start_m)}-{mon(end_m)} {y}"

        def months_in_window(start_m: int, end_m: int):
            return list(range(start_m, 13)) + list(range(1, end_m+1)) if start_m > end_m else list(range(start_m, end_m+1))

        def rolling_days_series(df, date_col, value_col, days, min_periods=1, out_region=None):
            """Return [date_col,'roll','region'] rolling-mean over a true days window; safe reset_index."""
            tmp = df[[date_col, value_col]].dropna().sort_values(date_col).copy()
            if tmp.empty:
                out = pd.DataFrame(columns=[date_col, "roll"])
                if out_region is not None: out["region"] = out_region
                return out
            tmp = tmp.set_index(pd.to_datetime(tmp[date_col])).drop(columns=[date_col])
            tmp.index.name = "_dt"
            tmp["roll"] = tmp[value_col].rolling(f"{days}D", min_periods=min_periods).mean()
            out = tmp.reset_index()[["_dt", "roll"]].rename(columns={"_dt": date_col})
            if out_region is not None: out = out.assign(region=out_region)
            return out

        def interpolate_to_daily_within_season(df, date_col, value_col, start_m, end_m, out_region=None):
            """
            Expand sparse monthly season-points to DAILY and interpolate ONLY inside season months.
            Off-season remains NaN -> Altair breaks the line (no diagonal bridges).
            """
            s = df[[date_col, value_col]].dropna().sort_values(date_col).copy()
            if s.empty:
                out = pd.DataFrame(columns=[date_col, "value"])
                if out_region is not None: out["region"] = out_region
                return out
            start, end = pd.to_datetime(s[date_col].min()), pd.to_datetime(s[date_col].max())
            days = pd.date_range(start, end, freq="D")
            ts = s.set_index(pd.to_datetime(s[date_col]))[value_col].reindex(days)
            mask = ts.index.month.isin(months_in_window(start_m, end_m))
            ts = ts.where(mask)
            ts = ts.interpolate(method="time", limit_area="inside")
            out = pd.DataFrame({date_col: ts.index, "value": ts.values}).dropna(subset=["value"])
            if out_region is not None: out["region"] = out_region
            return out

        # ---------------------------
        # Load + prep
        # ---------------------------
        df = load_table(file)
        if df is None:
            st.info("Upload a file to begin. Your table should have **one date column** and **numeric region columns**.")
            st.stop()

        df.columns = df.columns.map(lambda s: str(s).strip())

        def find_date_col(df: pd.DataFrame) -> str:
            for cand in ["date","Date","DATE","timestamp","Timestamp","datetime","Datetime"]:
                if cand in df.columns: return cand
            hit = [c for c in df.columns if "date" in str(c).lower()]
            if hit: return hit[0]
            dt = [c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])]
            if dt: return dt[0]
            raise ValueError("No date column found. Add a column named like 'Date'.")

        date_col = find_date_col(df)
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=False)
        if df[date_col].isna().all():
            st.error(f"Failed to parse any dates in '{date_col}'. Check raw values.")

        _df = df.copy()
        _df[date_col] = coerce_datetime(_df[date_col])
        _df = _df.sort_values(by=date_col)

        numeric_candidates = [c for c in _df.columns if c != date_col]
        for c in numeric_candidates:
            _df[c] = pd.to_numeric(_df[c], errors='coerce')

        with st.sidebar:
            value_cols = st.multiselect("Region columns", options=numeric_candidates,
                                        default=[c for c in numeric_candidates if not c.endswith("_LTA")])

        if not value_cols:
            st.warning("Select at least one region column.")
            st.stop()

        _df["season_year"] = _df[date_col].apply(lambda d: assign_season_year(d, start_month, end_month))
        filtered = _df.dropna(subset=["season_year"]).copy()
        if filtered.empty:
            st.error("No rows fall inside the selected season window. Try different months.")
            st.stop()

        filtered["season_label"] = filtered["season_year"].astype(int).apply(
            lambda y: season_label_from_year(y, start_month, end_month)
        )

        # Aggregations (seasonal)
        agg_method = st.radio("Aggregate by", ["Sum","Mean","Raw (daily)"], horizontal=True)
        _grouped = filtered[["season_year","season_label"]+value_cols].groupby(["season_year","season_label"], as_index=False)
        seasonal_sum  = _grouped.aggregate(np.nansum).sort_values(["season_year"])
        seasonal_mean = _grouped.aggregate(np.nanmean).sort_values(["season_year"])
        seasonal = seasonal_sum if agg_method=="Sum" else seasonal_mean
        agg_func = np.nansum if agg_method=="Sum" else np.nanmean

        # Seasons list
        season_options = (
            filtered[["season_year"]]
            .assign(season_label=lambda d: d["season_year"].astype(int).apply(lambda y: season_label_from_year(y,start_month,end_month)))
            .drop_duplicates().sort_values("season_year")
        )
        season_labels = season_options["season_label"].tolist()
        label_to_year = dict(zip(season_options["season_label"], season_options["season_year"]))

        st.subheader("📈 Seasons detected")
        st.dataframe(season_options[["season_label"]].reset_index(drop=True))

        base_regions = [c for c in value_cols if not c.endswith("_LTA")]

        st.markdown("---")
        view_mode = st.radio("View / Analysis", ["Global view","By Season","Summary (all regions)"])

        # ---------- chart helpers (no more schema error) ----------
        def enc_line(m, x_field: str, x_type: str, y_title: str):
            return (
                alt.Chart(m)
                .mark_line(interpolate="monotone", point=False, strokeWidth=2)
                .encode(
                    x=alt.X(x_field, type=x_type, title="Date" if x_type=="temporal" else "Season",
                            axis=alt.Axis(labelAngle=-30 if x_type=="temporal" else -35, labelOverlap=True, ticks=(x_type!="temporal"))),
                    y=alt.Y("value:Q", title=y_title),
                    color=alt.Color("region:N", title="Region", scale=alt.Scale(scheme="tableau10")),
                    tooltip=[
                        alt.Tooltip("region:N", title="Region"),
                        alt.Tooltip(x_field, type=x_type, title="Date" if x_type=="temporal" else "Season"),
                        alt.Tooltip("value:Q", title="Value", format=",.3f"),
                    ],
                )
            )

        def enc_bar(m, x_field: str, x_type: str, y_title: str):
            return (
                alt.Chart(m)
                .mark_bar()
                .encode(
                    x=alt.X(x_field, type=x_type, title="Date" if x_type=="temporal" else "Season",
                            axis=alt.Axis(labelAngle=-30 if x_type=="temporal" else -35, labelOverlap=True, ticks=(x_type!="temporal"))),
                    y=alt.Y("value:Q", title=y_title),
                    color=alt.Color("region:N", title="Region", scale=alt.Scale(scheme="tableau10")),
                    tooltip=[
                        alt.Tooltip("region:N", title="Region"),
                        alt.Tooltip(x_field, type=x_type, title="Date" if x_type=="temporal" else "Season"),
                        alt.Tooltip("value:Q", title="Value", format=",.3f"),
                    ],
                )
            )

        # ----------------------------
        # GLOBAL VIEW
        # ----------------------------
        if view_mode == "Global view":
            sel_regions = st.multiselect("Regions to plot", options=base_regions, default=base_regions)
            add_ma = st.checkbox("Add rolling average overlay", value=True)

            if agg_method == "Raw (daily)":
                interp_daily = st.checkbox("Interpolate to daily within season months", value=True)
                ma_win_days = st.slider("Rolling window (days)", 30, 365, 180, step=15)
            else:
                window_unit = st.radio("Overlay window", ["Seasons","Days"], horizontal=True)
                if window_unit=="Seasons":
                    ma_win_seasons = st.slider("Rolling window (seasons)", 2, 10, 5)
                else:
                    ma_win_days = st.slider("Rolling window (days)", 30, 365, 90, step=30)

            if not sel_regions:
                st.warning("Choose at least one region to plot.")
            else:
                left, right = st.columns(2)

                # ----- RAW (daily) -----
                if agg_method == "Raw (daily)":
                    if interp_daily:
                        pieces = [interpolate_to_daily_within_season(filtered, date_col, r, start_month, end_month, r) for r in sel_regions]
                        m = pd.concat(pieces, ignore_index=True) if pieces else pd.DataFrame(columns=[date_col,"value","region"])
                    else:
                        m = filtered[[date_col]+sel_regions].melt(id_vars=[date_col], var_name="region", value_name="value").sort_values(date_col)

                    # overlay (dashed)
                    if add_ma and not m.empty:
                        parts = []
                        for r in sel_regions:
                            df_reg = m[m["region"]==r][[date_col,"value"]].rename(columns={"value": r})
                            rolled = rolling_days_series(df_reg, date_col, r, ma_win_days, out_region=r)
                            parts.append(rolled)
                        ma_df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

                    # Line (left)
                    line_chart = enc_line(m, date_col, "temporal", "Value")
                    if add_ma and not ma_df.empty:
                        line_chart = alt.layer(
                            line_chart,
                            alt.Chart(ma_df).mark_line(strokeDash=[6,4], strokeWidth=2, opacity=0.9)
                                .encode(x=alt.X(date_col, type="temporal"), y="roll:Q", color="region:N")
                        )
                    left.altair_chart(line_chart.properties(title="Daily — Line"), use_container_width=True)

                    # Bar (right)
                    bar_chart = enc_bar(m, date_col, "temporal", "Value")
                    if add_ma and not ma_df.empty:
                        bar_chart = alt.layer(
                            bar_chart,
                            alt.Chart(ma_df).mark_line(strokeDash=[6,4], strokeWidth=2, opacity=0.9)
                                .encode(x=alt.X(date_col, type="temporal"), y="roll:Q", color="region:N")
                        )
                    right.altair_chart(bar_chart.properties(title="Daily — Bar"), use_container_width=True)

                # ----- Aggregated by season (Sum/Mean) -----
                else:
                    m = seasonal[["season_year","season_label"]+sel_regions] \
                            .melt(id_vars=["season_year","season_label"], var_name="region", value_name="value") \
                            .sort_values(["region","season_year"])

                    # overlay
                    ma_for_plot = pd.DataFrame()
                    if add_ma:
                        if window_unit=="Seasons":
                            m["roll"] = m.groupby("region", group_keys=False)["value"].apply(lambda s: s.rolling(ma_win_seasons, min_periods=1).mean())
                            ma_for_plot = m[["season_label","region","roll"]].dropna()
                        else:
                            parts = []
                            for r in sel_regions:
                                rolled = rolling_days_series(filtered[[date_col,"season_label",r]].dropna(), date_col, r, ma_win_days)
                                parts.append(rolled.groupby("season_label", as_index=False)["roll"].agg(agg_func).assign(region=r))
                            ma_for_plot = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

                    # Line (left)
                    line_chart = enc_line(m, "season_label", "nominal", f"{agg_method} over regions")
                    if add_ma and not ma_for_plot.empty:
                        line_chart = alt.layer(
                            line_chart,
                            alt.Chart(ma_for_plot).mark_line(strokeDash=[6,4], strokeWidth=2, opacity=0.9)
                                .encode(x=alt.X("season_label:N"), y="roll:Q", color="region:N")
                        )
                    left.altair_chart(line_chart.properties(title="By Season — Line"), use_container_width=True)

                    # Bar (right)
                    bar_chart = enc_bar(m, "season_label", "nominal", f"{agg_method} over regions")
                    if add_ma and not ma_for_plot.empty:
                        bar_chart = alt.layer(
                            bar_chart,
                            alt.Chart(ma_for_plot).mark_line(strokeDash=[6,4], strokeWidth=2, opacity=0.9)
                                .encode(x=alt.X("season_label:N"), y="roll:Q", color="region:N")
                        )
                    right.altair_chart(bar_chart.properties(title="By Season — Bar"), use_container_width=True)

        # ---------------------------
        # BY SEASON (date-level within one season)
        # ---------------------------
        elif view_mode == "By Season":
            sel_season = st.selectbox("Season", options=season_labels,
                                      index=len(season_labels)-1 if season_labels else 0)
            sel_regions = st.multiselect("Regions to plot", options=base_regions, default=base_regions)
            add_ma_season = st.checkbox("Add rolling average (days) in this season", value=False)
            if add_ma_season:
                ma_win_days_s = st.slider("Rolling window (days)", 7, 120, 30, step=7)

            if not sel_regions:
                st.warning("Choose at least one region to plot.")
            else:
                y = label_to_year[sel_season]
                within = filtered[filtered["season_year"]==y].copy()
                interp_here = st.checkbox("Interpolate to daily inside this season", value=True)

                if interp_here:
                    parts = [interpolate_to_daily_within_season(within, date_col, r, start_month, end_month, r) for r in sel_regions]
                    m = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=[date_col,"value","region"])
                else:
                    m = within[[date_col]+sel_regions].melt(id_vars=[date_col], var_name="region", value_name="value")

                if add_ma_season and not m.empty:
                    parts = []
                    for r in sel_regions:
                        df_reg = m[m["region"]==r][[date_col,"value"]].rename(columns={"value": r})
                        parts.append(rolling_days_series(df_reg, date_col, r, ma_win_days_s, out_region=r))
                    ma_df = pd.concat(parts, ignore_index=True)
                else:
                    ma_df = pd.DataFrame()

                left, right = st.columns(2)

                line_chart = enc_line(m, date_col, "temporal", "Value")
                if add_ma_season and not ma_df.empty:
                    line_chart = alt.layer(
                        line_chart,
                        alt.Chart(ma_df).mark_line(strokeDash=[6,4], strokeWidth=2, opacity=0.9)
                            .encode(x=alt.X(date_col, type="temporal"), y="roll:Q", color="region:N")
                    )
                left.altair_chart(line_chart.properties(title=f"{sel_season} — Line"), use_container_width=True)

                bar_chart = enc_bar(m, date_col, "temporal", "Value")
                if add_ma_season and not ma_df.empty:
                    bar_chart = alt.layer(
                        bar_chart,
                        alt.Chart(ma_df).mark_line(strokeDash=[6,4], strokeWidth=2, opacity=0.9)
                            .encode(x=alt.X(date_col, type="temporal"), y="roll:Q", color="region:N")
                    )
                right.altair_chart(bar_chart.properties(title=f"{sel_season} — Bar"), use_container_width=True)

        # ---------------------------
        # SUMMARY
        # ---------------------------
        elif view_mode == "Summary (all regions)":
            st.markdown("### Season × Region summary")
            pivot = seasonal[["season_label"]+base_regions].set_index("season_label").sort_index()
            st.dataframe(pivot)
            st.download_button("Download Season × Region (CSV)",
                pivot.reset_index().to_csv(index=False).encode("utf-8"),
                file_name="season_region_summary.csv", mime="text/csv")

            st.markdown("### Overall stats per region (across seasons)")
            stats = seasonal[base_regions].agg(["count","mean","std","min","max"]).T
            stats = stats.rename(columns={"count":"n_seasons","mean":f"mean_{'sum' if agg_method=='Sum' else 'mean'}"})
            st.dataframe(stats)
            st.download_button("Download Region Stats (CSV)",
                stats.reset_index().rename(columns={"index":"region"}).to_csv(index=False).encode("utf-8"),
                file_name="region_stats.csv", mime="text/csv")

        st.markdown("---")
        st.subheader("⬇️ Downloads")
        c1, c2 = st.columns(2)
        with c1:
            st.download_button("Download seasonal summary (CSV)",
                seasonal.to_csv(index=False).encode("utf-8"),
                file_name="seasonal_summary.csv", mime="text/csv")
        with c2:
            st.download_button("Download filtered rows (CSV)",
                filtered.to_csv(index=False).encode("utf-8"),
                file_name="filtered_rows.csv", mime="text/csv")

        st.caption("Tip: If the file is huge, pre-aggregate (daily→monthly) before upload to speed things up.")


    # ========= EDIT IIS DATA (no sidebar content) =========
    elif iis_option == "Edit IIS data":
        st.subheader("Edit IIS Data")
        iis_df = st.session_state.iis_df

        with st.expander("Add a column", expanded=True):
            new_col = st.text_input("Column name", placeholder="e.g., Portfolio Manager")
            col_type = st.selectbox("Data type", ["text","number","date"], index=0)
            default_val = None
            if col_type == "text":
                default_val = st.text_input("Default value (optional)")
            elif col_type == "number":
                default_val = st.number_input("Default value (optional)", value=0.0, step=1.0)
            else:
                default_val = st.date_input("Default value (optional)", value=None)
            if st.button("Add column"):
                if not new_col:
                    st.warning("Please enter a column name.")
                elif new_col in iis_df.columns:
                    st.warning(f"'{new_col}' already exists.")
                else:
                    if col_type == "date" and default_val is not None:
                        iis_df[new_col] = pd.to_datetime(default_val)
                    else:
                        iis_df[new_col] = default_val
                    st.session_state.iis_df = iis_df
                    st.success(f"Added column '{new_col}'.")

        with st.expander("Delete columns"):
            to_delete = st.multiselect("Select columns to delete", options=list(iis_df.columns))
            if st.button("Delete selected"):
                if not to_delete:
                    st.info("No columns selected.")
                else:
                    iis_df.drop(columns=to_delete, inplace=True, errors="ignore")
                    st.session_state.iis_df = iis_df
                    st.success(f"Deleted: {', '.join(to_delete)}")

        st.markdown("### Preview (editable cells)")
        edited = st.data_editor(st.session_state.iis_df, num_rows="dynamic", use_container_width=True)
        st.session_state.iis_df = edited

        st.markdown("### Save your edits (download)")
        csv_bytes = st.session_state.iis_df.to_csv(index=False).encode("utf-8")
        st.download_button("Download CSV", data=csv_bytes,
                           file_name="IIS_edited.csv", mime="text/csv")

        xbuf = io.BytesIO()
        with pd.ExcelWriter(xbuf, engine="xlsxwriter") as writer:
            st.session_state.iis_df.to_excel(writer, sheet_name=IIS_SHEET, index=False)
        st.download_button("Download Excel", data=xbuf.getvalue(),
                           file_name="IIS_edited.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.markdown("### Save back to main data source")
        st.caption("Overwrites the IIS sheet in the workbook after creating a timestamped backup.")
        confirm = st.checkbox("I understand this will replace the IIS sheet in the source file.")
        save_btn = st.button("Save IIS to workbook")
        if save_btn:
            if not confirm:
                st.warning("Please tick the confirmation box first.")
            else:
                try:
                    backup_path = backup_then_replace_iis_sheet(st.session_state.iis_df, DATA_PATH, IIS_SHEET)
                    st.success(f"Saved IIS sheet to '{DATA_PATH}'. Backup created: '{backup_path}'")
                    st.info("If the file is open in Excel/OneDrive lock, close it and try again.")
                    st.rerun()
                except PermissionError:
                    st.error("Permission denied. Is the workbook open or read-only?")
                except Exception as e:
                    st.error(f"Failed to write IIS sheet: {e}")




